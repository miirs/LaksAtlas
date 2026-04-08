import openpyxl

def clean(v):
    if v == '-' or v is None: return None
    try: return round(float(v), 2)
    except: return v

TARGET_YEARS = [2020, 2021, 2022, 2023]

def get_year_cols_row14(ws):
    row = list(ws.iter_rows(min_row=14, max_row=14, values_only=True))[0]
    yc = {}
    for ci, v in enumerate(row):
        if isinstance(v, int) and 2010 <= v <= 2030:
            yc[v] = ci
    return yc

def get_year_cols_row13(ws):
    row = list(ws.iter_rows(min_row=13, max_row=13, values_only=True))[0]
    yc = {}
    for ci, v in enumerate(row):
        if isinstance(v, (int, float)) and 2006 <= v <= 2030:
            yc[int(v)] = ci
    return yc

def extract_table(ws, year_cols, min_row, max_row, cols_per_year=4):
    data = {}
    for row in ws.iter_rows(min_row=min_row, max_row=max_row, values_only=True):
        region = row[0]
        if not region or not isinstance(region, str): continue
        if 'Totalt' in region or region.startswith('1)') or region.startswith('2)'): continue
        data[region] = {}
        for yr, ci in year_cols.items():
            vals = [clean(row[ci+i]) if ci+i < len(row) else None for i in range(cols_per_year)]
            data[region][yr] = vals
    return data

# LOAD
import os
os.chdir(r"C:\Users\Miiro\Projects\LaksAtlas\stats")

tap_wb  = openpyxl.load_workbook("sta-laks-mat-08-tap (1).xlsx", data_only=True)
kjop_wb = openpyxl.load_workbook("sta-laks-mat-05-kjop (2).xlsx", data_only=True)
lok_wb  = openpyxl.load_workbook("sta-laks-mat-02-lokaliteter (1).xlsx", data_only=True)
rens_wb = openpyxl.load_workbook("sta-laks-mat-10-rensefisk.xlsx", data_only=True)

# TAP
tap_ws  = tap_wb["Tap totalt"]
tap_yc  = get_year_cols_row14(tap_ws)
tap     = extract_table(tap_ws, tap_yc, 17, 26, 4)
tap_nat = {}
for row in tap_ws.iter_rows(min_row=26, max_row=26, values_only=True):
    for yr, ci in tap_yc.items():
        tap_nat[yr] = [clean(row[ci+i]) for i in range(4)]

# KJOP
kjop_ws = kjop_wb["Utsett"]
kjop_yc = get_year_cols_row14(kjop_ws)
kjop    = extract_table(kjop_ws, kjop_yc, 17, 26, 4)
kjop_nat= {}
for row in kjop_ws.iter_rows(min_row=26, max_row=26, values_only=True):
    for yr, ci in kjop_yc.items():
        kjop_nat[yr] = [clean(row[ci+i]) for i in range(4)]

# LOKALITETER
lok_ws  = lok_wb["Lokaliteter"]
lok_yc  = get_year_cols_row13(lok_ws)
lok     = {}
for row in lok_ws.iter_rows(min_row=16, max_row=26, values_only=True):
    region = row[0]
    if not region or not isinstance(region, str): continue
    if 'Totalt' in region or region.startswith('1)') or region.startswith('2)'): continue
    lok[region] = {}
    for yr, ci in lok_yc.items():
        v = row[ci]
        lok[region][yr] = int(v) if isinstance(v, (int, float)) and v is not None and float(v) == int(float(v)) else clean(v)
lok_nat = {}
for row in lok_ws.iter_rows(min_row=26, max_row=26, values_only=True):
    for yr, ci in lok_yc.items():
        v = row[ci]
        lok_nat[yr] = int(v) if isinstance(v, (int, float)) and v is not None and float(v) == int(float(v)) else clean(v)

# RENSEFISK
rens_ws = rens_wb["Fylke"]
rens_yc = {}
row14 = list(rens_ws.iter_rows(min_row=14, max_row=14, values_only=True))[0]
for ci, v in enumerate(row14):
    if isinstance(v, int) and 2010 <= v <= 2030:
        rens_yc[v] = ci
rens    = extract_table(rens_ws, rens_yc, 17, 24, 2)
rens_nat= {}
for row in rens_ws.iter_rows(min_row=24, max_row=24, values_only=True):
    for yr, ci in rens_yc.items():
        rens_nat[yr] = [clean(row[ci]), clean(row[ci+1])]

# =====================================================================
print("="*72)
print("LAKSATLAS REGIONAL DATA EXTRACTION (2020-2023)")
print("="*72)

print()
print("-"*72)
print("A. TAP (Deaths/Losses) -- 1000 individuals")
print("   Region                       | Year | Laks     | Regnbue  | Totalt")
print("-"*72)
for region in sorted(tap.keys()):
    for yr in TARGET_YEARS:
        if yr in tap[region]:
            v = tap[region][yr]
            print("  %-29s| %d | %9s | %8s | %9s" % (region, yr, str(v[0]), str(v[1]), str(v[3])))
print()
print("  NATIONAL TOTALS (TAP):")
for yr in TARGET_YEARS:
    if yr in tap_nat:
        v = tap_nat[yr]
        print("  %d: Laks=%-12s Regnbue=%-10s Orret=%-8s Totalt=%-10s" % (yr, str(v[0]), str(v[1]), str(v[2]), str(v[3])))

print()
print("-"*72)
print("B. KJOP/UTSETT (Smolt Put-In) -- 1000 individuals")
print("   Region                       | Year | Laks     | Regnbue  | Totalt")
print("-"*72)
for region in sorted(kjop.keys()):
    for yr in TARGET_YEARS:
        if yr in kjop[region]:
            v = kjop[region][yr]
            print("  %-29s| %d | %9s | %8s | %9s" % (region, yr, str(v[0]), str(v[1]), str(v[3])))
print()
print("  NATIONAL TOTALS (KJOP):")
for yr in TARGET_YEARS:
    if yr in kjop_nat:
        v = kjop_nat[yr]
        print("  %d: Laks=%-12s Regnbue=%-10s Totalt=%-10s" % (yr, str(v[0]), str(v[1]), str(v[3])))

print()
print("-"*72)
print("C. LOKALITETER (Active Sea Sites at Dec 31) -- count")
print("   Region                       | 2020 | 2021 | 2022 | 2023")
print("-"*72)
for region in sorted(lok.keys()):
    vals = [str(lok[region].get(yr, 'N/A')) for yr in TARGET_YEARS]
    print("  %-29s| %5s| %5s| %5s| %5s" % (region, vals[0], vals[1], vals[2], vals[3]))
nat = [str(lok_nat.get(yr, 'N/A')) for yr in TARGET_YEARS]
print("  %-29s| %5s| %5s| %5s| %5s" % ("NATIONAL TOTAL", nat[0], nat[1], nat[2], nat[3]))

print()
print("-"*72)
print("D. RENSEFISK (Cleaner Fish Deployed) -- 1000 fish / 1000 NOK")
print("   Region                       | Year | Antall(k)| Verdi(kNOK)")
print("-"*72)
for region in sorted(rens.keys()):
    for yr in TARGET_YEARS:
        if yr in rens[region]:
            v = rens[region][yr]
            print("  %-29s| %d | %9s | %11s" % (region, yr, str(v[0]), str(v[1])))
print()
print("  NATIONAL TOTALS (RENSEFISK):")
for yr in TARGET_YEARS:
    if yr in rens_nat:
        v = rens_nat[yr]
        print("  %d: Antall=%-12s Verdi=%-12s" % (yr, str(v[0]), str(v[1])))

print()
print("="*72)
print("NOTES:")
print("  1. 2020-2022: Troms+Finnmark merged as 'Troms og Finnmark' fylke")
print("     2023+: split back to 'Finnmark' and 'Troms' separately")
print("  2. Rensefisk uses 'Finnmark og Troms' for all years (always combined)")
print("  3. None = '-' in source (not applicable / not reported)")
print("  4. All counts are x1000 individuals; Verdi is x1000 NOK")
print("  5. Lokaliteter: Agder not in main salmon regions but included")
print("="*72)
